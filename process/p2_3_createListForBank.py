import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from typing import Tuple, Dict
from process.partner_resolution import resolve_partner_columns
from process.transaction_details import (
    build_direct_sales_details,
    consolidate_partner_aliases,
    resolve_payment_partner_name,
)
from process.capital_movement import build_capital_movement_list

VALID_TRANSACTION_NULL_STRINGS = {"", "nan", "none", "null", "<na>"}


def has_valid_transaction_no(value) -> bool:
    if pd.isna(value):
        return False
    return str(value).strip().lower() not in VALID_TRANSACTION_NULL_STRINGS


def is_opening_entry(row: pd.Series) -> bool:
    opening_keywords = ('開始仕訳', '期首', '前期繰越', '繰越')
    fields = (
        row.get('description'), row.get('partner'), row.get('debit_partner'), row.get('credit_partner'),
        row.get('debit_account'), row.get('credit_account')
    )
    return any(
        keyword in str(value)
        for value in fields if pd.notna(value)
        for keyword in opening_keywords
    )


def exclude_opening_cash_movements(df: pd.DataFrame) -> pd.DataFrame:
    """1年目・2年目の期首にある開始／繰越仕訳の現預金増減をゼロにする。"""
    result = df.copy()
    min_date = result['date'].min()
    if pd.isna(min_date):
        return result

    first_opening_date = pd.Timestamp(min_date.year, min_date.month, 1)
    second_opening_date = first_opening_date + pd.DateOffset(years=1)
    is_target_opening_date = result['date'].isin([first_opening_date, second_opening_date])
    opening_row = result.apply(is_opening_entry, axis=1)
    opening_transaction_nos = set(
        result.loc[is_target_opening_date & opening_row, 'transaction_no']
            .dropna()
            .loc[lambda s: s.apply(has_valid_transaction_no)]
    )
    same_opening_transaction = result['transaction_no'].isin(opening_transaction_nos)
    is_opening_journal = is_target_opening_date & (opening_row | same_opening_transaction)
    result.loc[is_opening_journal, 'cash_diff'] = 0.0
    return result

# ==========================================================
# 「資金移動用途推定」シート専用の列幅設定（文字数換算）
# 後から列幅を変更したい場合は、以下の数値を編集してください。
# ==========================================================
BANK_LIST_COLUMN_WIDTHS = {
    "資金移動日": 16,
    "移動金額": 18,
    "用途充当額": 18,
    "金額差（未充当額）": 20,
    "金額差率（未充当率）": 18,
    "想定用途": 20,
    "資金移動仕訳": 48,
    "用途推定仕訳": 58,
    "一致区分": 15,
    "信頼度": 10
}

def format_journal_entry(account, sub_account, partner, amount) -> str:
    """
    関連仕訳の文字列をフォーマットするヘルパー関数。
    形式: 【[勘定科目]】[補助科目]（[金額]円）
    ※補助科目が空（NaNまたは空文字）の場合、代わりに摘要（partner）を使用します。
    ※括弧は全角「（」および「）」を使用します。
    """
    acc_str = str(account).strip() if pd.notna(account) else ""
    sub_str = str(sub_account).strip() if pd.notna(sub_account) else ""
    pat_str = str(partner).strip() if pd.notna(partner) else ""
    
    # 補助科目が無い場合は摘要を使用
    middle_str = sub_str if sub_str != "" else pat_str
    
    amt_str = f"{amount:.0f}" if pd.notna(amount) else "0"
    
    return f"【{acc_str}】{middle_str}（{amt_str}円）"


def build_sales_receipt_list(df_receipts: pd.DataFrame, journal: pd.DataFrame) -> pd.DataFrame:
    """売掛金回収について、債権回収額・実入金額・手数料を分けて表示する。"""
    columns = ["日付", "金額", "売掛金回収額", "実入金額", "差額", "振込手数料",
               "対応状態", "相手科目(借方/貸方)", "摘要", "貸方補助科目"]
    if df_receipts.empty:
        return pd.DataFrame(columns=columns)

    fee_mask = journal["debit_account"].fillna("").astype(str).str.contains("支払手数料", na=False)
    fees = journal[fee_mask].copy()
    fees["_fee_amount"] = pd.to_numeric(fees["debit_amount"], errors="coerce").fillna(0.0)
    used_fee_indices = set()
    records = []

    for _, row in df_receipts.iterrows():
        gross_value = pd.to_numeric(row.get("credit_amount"), errors="coerce")
        net_value = pd.to_numeric(row.get("debit_amount"), errors="coerce")
        gross = 0.0 if pd.isna(gross_value) else float(gross_value)
        net = 0.0 if pd.isna(net_value) else float(net_value)
        difference = max(gross - net, 0.0)
        matched_fee = 0.0

        transaction_no = row.get("transaction_no")
        if has_valid_transaction_no(transaction_no) and difference > 0:
            candidates = fees[
                (fees["transaction_no"] == transaction_no)
                & (~fees.index.isin(used_fee_indices))
                & ((fees["_fee_amount"] - difference).abs() <= 1.0)
            ]
            if not candidates.empty:
                fee_index = candidates.index[0]
                matched_fee = float(candidates.loc[fee_index, "_fee_amount"])
                used_fee_indices.add(fee_index)

        status = "一致" if difference == 0 else ("手数料一致" if matched_fee > 0 else "差額要確認")
        partner = row.get("ar_partner")
        if pd.isna(partner):
            partner = row.get("credit_partner")
        records.append({
            "日付": row.get("date"), "金額": net, "売掛金回収額": gross,
            "実入金額": net, "差額": gross - net, "振込手数料": matched_fee,
            "対応状態": status, "相手科目(借方/貸方)": row.get("credit_account"),
            "摘要": row.get("description"), "貸方補助科目": partner,
        })
    return pd.DataFrame(records, columns=columns).sort_values("日付")

def cleanse_journal(df: pd.DataFrame) -> pd.DataFrame:
    """
    仕訳データのクレンジング処理。
    借方金額または貸方金額にマイナス値がある場合、それを正の数に変換し、
    debit/creditを反転させるクレンジングを行う。
    """
    df_clean = df.copy()
    
    # 借方金額がマイナスの場合の処理
    debit_minus = df_clean['debit_amount'] < 0
    if debit_minus.any():
        for idx in df_clean[debit_minus].index:
            row = df_clean.loc[idx]
            df_clean.loc[idx, ['debit_amount', 'credit_amount', 'debit_account', 'credit_account', 'debit_partner', 'credit_partner']] = \
                [0.0, abs(row['debit_amount']), row['credit_account'], row['debit_account'], row['credit_partner'], row['debit_partner']]
                
    # 貸方金額がマイナスの場合の処理
    credit_minus = df_clean['credit_amount'] < 0
    if credit_minus.any():
        for idx in df_clean[credit_minus].index:
            row = df_clean.loc[idx]
            df_clean.loc[idx, ['debit_amount', 'credit_amount', 'debit_account', 'credit_account', 'debit_partner', 'credit_partner']] = \
                [abs(row['credit_amount']), 0.0, row['credit_account'], row['debit_account'], row['credit_partner'], row['debit_partner']]
                
    return df_clean


def select_long_ar_fill_key(origin_value, evaluation_value):
    """長期未回収売掛の行色キー。発生原因による灰色を評価色より優先する。"""
    origin = "" if pd.isna(origin_value) else str(origin_value)
    evaluation = "" if pd.isna(evaluation_value) else str(evaluation_value)
    if origin == "預金支出起点・要確認":
        return "grey"
    if "🔴" in evaluation or "長期滞留" in evaluation:
        return "red"
    if "🔶" in evaluation or "要整理" in evaluation:
        return "orange"
    if "⚠️" in evaluation or "注意" in evaluation:
        return "yellow"
    return None

def create_bank_excel(df_journal: pd.DataFrame, df_bs: pd.DataFrame) -> Tuple[bytes, Dict]:
    """
    銀行説明用リストのエクセルブック（全8シート）を作成する。
    また、売上計上思想指数の算出データを辞書形式で返す。
    """
    # 1. 仕訳クレンジング
    # 負額仕訳の借貸反転後は、標準化時の解決済み列を使い回さず一度だけ再解決する。
    df_j = resolve_partner_columns(cleanse_journal(df_journal), force=True)
    df_j['date'] = pd.to_datetime(df_j['date'], errors='coerce')
    df_j = df_j.dropna(subset=['date']).sort_values('date')
    
    # --- 指標14: 売上計上思想指数 ---
    is_sales_debit = df_j['debit_account'].str.contains('売上', na=False) & ~df_j['debit_account'].str.contains('雑収入', na=False)
    is_sales_credit = df_j['credit_account'].str.contains('売上', na=False) & ~df_j['credit_account'].str.contains('雑収入', na=False)
    df_sales = df_j[is_sales_debit | is_sales_credit].copy()
    
    a_keywords = ['概算', '見込', '仮', '仮売上', '予想']
    b_keywords = ['修正', '取消', '訂正', '振替', '再計算']
    all_keywords = a_keywords + b_keywords
    
    def contains_keywords(val):
        if pd.isna(val):
            return False
        val_str = str(val)
        return any(kw in val_str for kw in all_keywords)
        
    # 空Seriesでもboolean maskとして扱えるようdtypeを固定する。
    # dtype=objectの空Seriesを df[series] に渡すと列選択と解釈され、
    # 後続で debit_amount / credit_amount が消えるため。
    df_sales['is_target'] = df_sales['description'].apply(contains_keywords).astype(bool)
    df_sales_target = df_sales.loc[df_sales['is_target']].copy()
    
    total_sales_count = len(df_sales)
    target_sales_count = len(df_sales_target)
    sales_index = (target_sales_count / total_sales_count * 100) if total_sales_count > 0 else 0.0
    
    sales_index_data = {
        "index": sales_index,
        "target_count": target_sales_count,
        "total_count": total_sales_count
    }
    
    # 1. 売上計上思想_該当仕訳
    df_sheet1 = df_sales_target.copy()
    # 金額は最大値を取得（クレンジングで正の数に変換済み）
    df_sheet1['amount'] = df_sheet1[['debit_amount', 'credit_amount']].max(axis=1)
    df_sheet1 = df_sheet1[['date', 'amount', 'description', 'debit_account', 'credit_account', 'credit_partner']].rename(columns={'description': '摘要', 'credit_partner': '貸方補助科目'})
    df_sheet1 = df_sheet1.sort_values('date')
    
    # 2. 売上計上思想_全売上仕訳
    df_sheet2 = df_sales.copy()
    df_sheet2['amount'] = df_sheet2[['debit_amount', 'credit_amount']].max(axis=1)
    df_sheet2 = df_sheet2[['date', 'amount', 'description', 'debit_account', 'credit_account', 'credit_partner']].rename(columns={'description': '摘要', 'credit_partner': '貸方補助科目'})
    df_sheet2 = df_sheet2.sort_values('date')
    
    # --- 指標15: 売上入金・直入金売上リスト ---
    debit_pat = '預金|現金|受取手形|電信|当座|普通'
    credit_pat = '売掛|未収|買入金銭債権'
    
    # 3. 売上入金
    df_nyukin = df_j[
        df_j['debit_account'].str.contains(debit_pat, na=False) &
        df_j['credit_account'].str.contains(credit_pat, na=False)
    ].copy()
    df_sheet3 = build_sales_receipt_list(df_nyukin, df_j)
        
    # 4. 直入金売上
    df_choku = build_direct_sales_details(df_j)
    df_sheet4 = pd.DataFrame()
    if not df_choku.empty:
        df_sheet4['日付'] = df_choku['date']
        df_sheet4['金額'] = df_choku['amount']
        df_sheet4['相手科目(借方/貸方)'] = df_choku['credit_account']
        df_sheet4['摘要'] = df_choku['description']
        df_sheet4['取引先'] = df_choku['partner']
        df_sheet4['取引先取得元'] = df_choku['source']
        df_sheet4 = df_sheet4.sort_values('日付')
        
    # --- 指標16: 直払いリスト ---
    exclude_pat = '買掛|未払|借入|利息|税|仮払|手数料'
    is_credit_yokin = df_j['credit_account'].str.contains('普通預金|当座預金', na=False)
    is_not_excluded = ~df_j['debit_account'].str.contains(exclude_pat, na=False)
    df_pay_base = df_j[is_credit_yokin & is_not_excluded].copy()
    
    def get_pay_category(row):
        acc = f"{row.get('debit_account', '')} {row.get('debit_partner', '')}"
        if '仕入' in acc:
            return '仕入支払'
        elif '外注' in acc:
            return '外注支払'
        elif any(k in acc for k in ['地代', '家賃', '賃借料']):
            return '固定費支払'
        elif any(k in acc for k in ['広告', '宣伝']):
            return '販促費支払'
        elif any(k in acc for k in ['消耗品', '手数料', '修繕', '運賃', '通信', '水道', '光熱', '電気', 'ガス']):
            return 'その他支払'
        return None
        
    df_pay_base['category'] = df_pay_base.apply(get_pay_category, axis=1)
    df_pay = df_pay_base[df_pay_base['category'].notna()].copy()
    
    df_sheet5 = pd.DataFrame()
    if not df_pay.empty:
        df_pay['支払先'] = df_pay.apply(resolve_payment_partner_name, axis=1)
        df_pay['支払先'] = consolidate_partner_aliases(df_pay['支払先']).fillna('取引先不明')
        df_sheet5['日付'] = df_pay['date']
        df_sheet5['金額'] = df_pay['credit_amount']
        df_sheet5['借方科目'] = df_pay['debit_account']
        df_sheet5['摘要'] = df_pay['description']
        df_sheet5['借方補助科目'] = df_pay['debit_partner']
        df_sheet5['支払先'] = df_pay['支払先']
        df_sheet5['カテゴリ'] = df_pay['category']
        df_sheet5 = df_sheet5.sort_values('日付')
        
    # --- 指標17: 預金体力推移 ---
    bs_available = (not df_bs.empty) and ("期末現預金合計" in df_bs.columns)
    df_sheet6 = pd.DataFrame()
    
    if bs_available:
        cash_balance_end = df_bs["期末現預金合計"].iloc[0]
        
        # 現預金科目の増減計算
        # 借方に現預金科目がある場合はプラス、貸方にある場合はマイナス
        cash_pat = '預金|現金|当座|普通'
        df_j['debit_is_cash'] = df_j['debit_account'].str.contains(cash_pat, na=False)
        df_j['credit_is_cash'] = df_j['credit_account'].str.contains(cash_pat, na=False)
        
        df_j['cash_diff'] = df_j.apply(
            lambda r: (r['debit_amount'] if r['debit_is_cash'] else 0) - (r['credit_amount'] if r['credit_is_cash'] else 0),
            axis=1
        )
        
        # 1年目・2年目の期首日に計上された開始／繰越仕訳だけを除外する。
        # 同じ取引Noの複合仕訳は、いずれかの行に開始／繰越表示があれば一括で除外する。
        df_j = exclude_opening_cash_movements(df_j)
        
        # 日次で集計
        df_daily = df_j.groupby('date')['cash_diff'].sum().reset_index()
        
        # 最終日から遡って残高を計算
        df_daily = df_daily.sort_values('date', ascending=False).reset_index(drop=True)
        daily_balances = []
        current_bal = cash_balance_end
        
        for idx, row in df_daily.iterrows():
            daily_balances.append({
                "date": row['date'],
                "balance": current_bal
            })
            current_bal -= row['cash_diff']
            
        df_bal = pd.DataFrame(daily_balances).sort_values('date').reset_index(drop=True)
        
        # 月次集計
        df_bal['year_month'] = df_bal['date'].dt.to_period('M')
        
        monthly_data = []
        for ym, group in df_bal.groupby('year_month'):
            # 月末残高
            end_bal = group.sort_values('date').iloc[-1]['balance']
            # 月内最低残高
            min_row = group.sort_values('balance').iloc[0]
            min_bal = min_row['balance']
            min_date = min_row['date']
            
            monthly_data.append({
                "対象年月": ym.strftime("%y%m"),
                "月末残高": int(end_bal),
                "月内最低残高": int(min_bal),
                "月内最低残高の記録日": min_date
            })
            
        df_sheet6 = pd.DataFrame(monthly_data).sort_values("対象年月")
        
    # 資金移動用途推定は、取引No単位・諸口展開・移動先口座追跡を行う専用ロジックで確定する。
    df_sheet7 = build_capital_movement_list(df_j)

    # --- 指標19: 長期未回収売掛リスト ---
    # 期間と期首日の算出
    min_date = df_j['date'].min()
    start_date = pd.Timestamp(year=min_date.year, month=min_date.month, day=1)
    
    # 2年目の期首日（1年目のちょうど1年後）を判定するための年月
    year2_year = start_date.year + 1
    year2_month = start_date.month
    
    # 2年目の期首日にある「借方が売掛金系」の仕訳をすべて除外するフィルター
    # （開始仕訳等のキーワードに頼らず、該当日の借方発生を一律で除外する）
    ar_pattern_for_carryover = '売掛|未収|買入金銭債権'
    is_year2_opening_debit = (
        (df_j['date'].dt.year == year2_year) & 
        (df_j['date'].dt.month == year2_month) & 
        (df_j['date'].dt.day == 1) & 
        df_j['debit_account'].astype(str).str.contains(ar_pattern_for_carryover, na=False)
    )
    
    clean_df = df_j[~is_year2_opening_debit].copy()
    
    import re
    
    def do_cleanse(val):
        if pd.isna(val):
            return ''
        s = str(val).strip()
        if not s or s.lower() == 'nan':
            return ''
            
        # B: 法人格の削除（スペース関係なく除去）
        corp_patterns = [
            r"株式会社", r"有限会社", r"合資会社", r"合名会社", r"合同会社",
            r"\(株\)", r"（株）", r"\(有\)", r"（有）", r"\(合\)", r"（合）",
            r"㈱", r"㈲", r"㈴", r"㈵", r"法人",
            r"カブシキガイシャ", r"ユウゲンガイシャ", r"ゴウドウガイシャ",
            r"\(カ\)", r"（カ）", r"カ\)", r"\(カ", r"（カ", r"カ）",
            r"カ\.", r"\.カ",
            r"\(ユ\)", r"（ユ）", r"ユ\)", r"\(ユ\)", r"（ユ", r"ユ）",
            r"ユ\.", r"\.ユ",
            r"トクヒ\)", r"\(トクヒ", r"トクヒ"
        ]
        for pat in corp_patterns:
            s = re.sub(pat, "", s)
            
        # A: 「(」または「（」以降を削除
        s = re.split(r'[(（]', s)[0]
        
        # C, D: スペースで分割してトークンごとに判定
        tokens = re.split(r'[ 　]+', s)
        filtered_tokens = []
        for t in tokens:
            if not t:
                continue
            # C: 特定キーワードが存在するトークンを除外
            if any(k in t for k in ['銀行', '金庫', '信用組合', '信組', '農協', '営業部', '支店', '預金', '振込', 'フリコミ', 'ﾌﾘｺﾐ']):
                continue
            # D: 6桁以上の数字を含むトークンを除外（口座番号や管理番号対策）
            if re.search(r'\d{6,}', t):
                continue
            filtered_tokens.append(t)
            
        # 最後にスペースなしで結合
        s = "".join(filtered_tokens)
        s = re.sub(r"^[.\-_ー]+|[.\-_ー]+$", "", s)
        return s

    def get_clean_partner_for_ar(row):
        return do_cleanse(row.get('ar_partner'))

    # 名寄せ用の取引先名
    clean_df['partner_clean'] = clean_df.apply(get_clean_partner_for_ar, axis=1)
    
    # 発生・回収のどちらにおいても、補助科目と摘要がどちらも空欄（追跡不可）のものは除外
    clean_df = clean_df[clean_df['partner_clean'] != '']
    
    
    # 【包含一致 名寄せロジック】
    # 短い名称が長い名称に含まれる場合、同一取引先とみなして名寄せする。
    protected_individuals = set(
        clean_df.loc[clean_df.get('ar_partner_kind', pd.Series(index=clean_df.index, dtype='object')) == 'individual', 'partner_clean']
        .dropna().astype(str)
    )
    clean_df['partner_clean'] = consolidate_partner_aliases(
        clean_df['partner_clean'], protected_names=protected_individuals
    ).fillna('')
        
    # AR（売掛・未収）系科目の判定パターン
    ar_pattern = '売掛|未収|買入金銭債権'
    
    # 期首残高仕訳（真の期首日の借方売掛金/未収入金/買入金銭債権）
    is_opening = (clean_df['date'] == start_date) & clean_df['debit_account'].str.contains(ar_pattern, na=False)
    opening_df = clean_df[is_opening]
    opening_bal = opening_df.groupby('partner_clean')['debit_amount'].sum().to_dict()
    
    # 借方・貸方それぞれのAR判定
    is_debit_ar = clean_df['debit_account'].str.contains(ar_pattern, na=False)
    is_credit_ar = clean_df['credit_account'].str.contains(ar_pattern, na=False)
    
    # 【追加フィルター】同一行内で借方・貸方の両方がAR系の場合は「請求締め」等の内部振替とみなし除外
    is_internal_ar_transfer = is_debit_ar & is_credit_ar
    
    # 期中発生仕訳（借方AR、期首仕訳および内部振替を除く）
    df_gen = clean_df[is_debit_ar & ~is_opening & ~is_internal_ar_transfer].copy()

    # 発生原因は残高計算から分離して表示用に分類する。
    valid_tx = clean_df['transaction_no'].notna() & ~clean_df['transaction_no'].astype(str).str.lower().isin(
        ['', 'nan', 'none', 'null', '<na>']
    )
    sales_row = clean_df['credit_account'].astype(str).str.contains('売上', na=False)
    sales_transactions = set(clean_df.loc[valid_tx & sales_row, 'transaction_no'].astype(str))

    def classify_ar_origin(row):
        credit_account = str(row.get('credit_account', ''))
        if '売上' in credit_account:
            return '売上起点'
        tx = row.get('transaction_no')
        if pd.notna(tx) and str(tx) in sales_transactions:
            return '売上起点（複合仕訳）'
        if re.search(r'預金|現金|当座|普通|定期|別段|手形|電信', credit_account):
            return '預金支出起点・要確認'
        return 'その他・要確認'

    if not df_gen.empty:
        df_gen['ar_origin'] = df_gen.apply(classify_ar_origin, axis=1)
    
    # 期中回収仕訳（貸方AR、内部振替を除く）
    df_kai = clean_df[is_credit_ar & ~is_internal_ar_transfer].copy()
    
    base_date = clean_df['date'].max()
    uncollected_items = []
    debug_ar_list = []
    
    # 全取引先の一覧
    all_partners = set(clean_df['partner_clean'].unique())
    
    for p in all_partners:
        op = opening_bal.get(p, 0.0)
        deb_sum = df_gen[df_gen['partner_clean'] == p]['debit_amount'].sum()
        cred_sum = df_kai[df_kai['partner_clean'] == p]['credit_amount'].sum()
        
        # 逆算フォールバック：回収額が発生と期首を上回る場合は期首を補正
        op_adj = max(op, cred_sum - deb_sum)
        
        # 期内残存売掛金残高（期末残高）
        rem_bal = op_adj + deb_sum - cred_sum
        
        # デバッグ用レコードの追加
        debug_ar_list.append({
            "取引先(名寄せ後)": p,
            "期首残高": op,
            "期中発生(借方)": deb_sum,
            "期中回収(貸方)": cred_sum,
            "期首補正額": op_adj - op,
            "期末残高": rem_bal
        })
        
        if rem_bal <= 0:
            continue
            
        # 発生仕訳（借方）を古い順（昇順）に整理する。
        # まず、期首残高（補正後）を start_date (2023-09-01) の仮想的な発生レコードとしてリストの先頭に配置する。
        p_gens_list = []
        if op_adj > 0:
            # 代表的な科目名を特定
            p_gens_temp = clean_df[(clean_df['partner_clean'] == p) & is_debit_ar]
            acc_name = p_gens_temp['debit_account'].iloc[0] if not p_gens_temp.empty else "売掛金"
            p_gens_list.append({
                "date": start_date,
                "debit_amount": op_adj,
                "debit_account": acc_name,
                "ar_origin": "期首残高"
            })
            
        # 次に、期中発生（期首日以外の借方仕訳）を古い順（昇順）に追加する
        p_gen_mid = df_gen[df_gen['partner_clean'] == p].sort_values('date', ascending=True)
        for _, row in p_gen_mid.iterrows():
            p_gens_list.append({
                "date": row['date'],
                "debit_amount": row['debit_amount'],
                "debit_account": row['debit_account'],
                "ar_origin": row['ar_origin']
            })
            
        # 回収の総額 `cred_sum` を用いて、古い発生から順次消し込む
        remaining_cred = cred_sum
        uncollected_gens = []
        
        for g in p_gens_list:
            g_amt = g['debit_amount']
            if remaining_cred >= g_amt:
                # この発生は完全に回収された
                remaining_cred -= g_amt
            elif remaining_cred > 0:
                # この発生は部分的に回収された
                uncollected_amt = g_amt - remaining_cred
                remaining_cred = 0.0
                uncollected_gens.append({
                    "date": g['date'],
                    "amount": uncollected_amt,
                    "debit_account": g['debit_account'],
                    "ar_origin": g['ar_origin']
                })
            else:
                # 回収原資が尽きているため、この発生は丸ごと未回収
                uncollected_gens.append({
                    "date": g['date'],
                    "amount": g_amt,
                    "debit_account": g['debit_account'],
                    "ar_origin": g['ar_origin']
                })
                
        # 未回収明細のうち、31日以上滞留しているものをリストアップする
        for ug in uncollected_gens:
            gen_date = ug['date']
            amt = ug['amount']
            if amt <= 0.01:
                continue
                
            days = (base_date - gen_date).days
            if days >= 31:
                if days >= 91:
                    eval_str = "🔴長期滞留"
                elif days >= 61:
                    eval_str = "🔶要整理"
                else:
                    eval_str = "⚠️注意"
                    
                uncollected_items.append({
                    "発生日": gen_date,
                    "滞留日数": days,
                    "金額": amt,
                    "取引先": p,
                    "勘定科目": ug['debit_account'],
                    "発生原因": ug['ar_origin'],
                    "評価": eval_str
                })
                
    df_sheet8 = pd.DataFrame(uncollected_items, columns=["発生日", "滞留日数", "金額", "取引先", "勘定科目", "発生原因", "評価"])
    if not df_sheet8.empty:
        df_sheet8 = df_sheet8.sort_values("滞留日数", ascending=False)
        
    df_debug_ar = pd.DataFrame(debug_ar_list)
    if not df_debug_ar.empty:
        df_debug_ar = df_debug_ar.sort_values("期末残高", ascending=False)
        
    # --- Excelの出力・装飾 (openpyxl) ---
    wb = openpyxl.Workbook()
    # デフォルトのSheetを削除するため、まずはシート作成を行ってから最後に削除する
    
    sheets_info = [
        ("売上計上思想_該当仕訳", df_sheet1, ["日付", "金額", "摘要", "借方科目", "貸方科目", "貸方補助科目"]),
        ("売上計上思想_全売上仕訳", df_sheet2, ["日付", "金額", "摘要", "借方科目", "貸方科目", "貸方補助科目"]),
        ("売上入金", df_sheet3, ["日付", "金額", "売掛金回収額", "実入金額", "差額", "振込手数料", "対応状態", "相手科目(借方/貸方)", "摘要", "貸方補助科目"]),
        ("直入金売上", df_sheet4, ["日付", "金額", "相手科目(借方/貸方)", "摘要", "取引先", "取引先取得元"]),
        ("直払いリスト", df_sheet5, ["日付", "金額", "借方科目", "摘要", "借方補助科目", "支払先", "カテゴリ"]),
        ("預金体力推移", df_sheet6, ["対象年月", "月末残高", "月内最低残高", "月内最低残高の記録日"]),
        ("資金移動用途推定", df_sheet7, ["資金移動日", "移動金額", "用途充当額", "金額差（未充当額）", "金額差率（未充当率）", "想定用途", "資金移動仕訳", "用途推定仕訳", "一致区分", "信頼度"]),
        ("長期未回収売掛", df_sheet8, ["発生日", "滞留日数", "金額", "取引先", "勘定科目", "発生原因", "評価"])
    ]
    
    # スタイル定義
    FONT_NAME = "BIZ UDゴシック"
    font_regular = Font(name=FONT_NAME, size=10)
    font_bold = Font(name=FONT_NAME, size=10, bold=True)
    font_header = Font(name=FONT_NAME, size=11, color="FFFFFF", bold=True)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_red = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    fill_orange = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    fill_grey = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Side(border_style="thin", color="D3D3D3")
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    for sheet_name, df_data, cols in sheets_info:
        ws = wb.create_sheet(title=sheet_name)
        
        # 1. 預金体力推移でB/Sデータがない場合
        if sheet_name == "預金体力推移" and not bs_available:
            ws.cell(row=2, column=2, value="※貸借対照表（B/S）がアップロードされていないため、預金体力推移は算出できません。").font = font_bold
            ws.column_dimensions['B'].width = 80
            continue
            
        # ヘッダー書き込み
        ws.row_dimensions[1].height = 28
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
            
        # データ書き込み
        if not df_data.empty:
            for row_idx, row_data in enumerate(df_data.values, 2):
                row_height = 20
                
                # 「資金移動用途推定」シートはセル内改行数に応じて行の高さを広げる
                if sheet_name == "資金移動用途推定":
                    idx_transfer = cols.index("資金移動仕訳")
                    idx_purpose = cols.index("用途推定仕訳")
                    val_transfer = str(row_data[idx_transfer]) if pd.notna(row_data[idx_transfer]) else ""
                    val_purpose = str(row_data[idx_purpose]) if pd.notna(row_data[idx_purpose]) else ""
                    lines_transfer = val_transfer.count('\n') + 1
                    lines_purpose = val_purpose.count('\n') + 1
                    max_lines = max(lines_transfer, lines_purpose)
                    row_height = max(20, max_lines * 16) # 1行あたり16pt
                    
                ws.row_dimensions[row_idx].height = row_height
                
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_regular
                    cell.border = border_cell
                    
                    # 型に応じたフォーマットと配置
                    if isinstance(val, pd.Timestamp):
                        cell.value = val.strftime('%Y-%m-%d')
                        cell.alignment = align_center
                    elif isinstance(val, (int, float, np.integer, np.floating)):
                        cell.value = val
                        col_name = cols[col_idx-1]
                        if col_name == "金額差率（未充当率）":
                            cell.number_format = "0.0%"
                            cell.alignment = align_right
                        elif "金額" in col_name or col_name == "月末残高" or col_name == "月内最低残高":
                            cell.number_format = "#,##0"
                            cell.alignment = align_right
                        elif col_name == "滞留日数":
                            cell.number_format = "#,##0"
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_left
                    else:
                        cell.value = str(val) if pd.notna(val) else ""
                        cell.alignment = align_left
                        
                    # 資金移動用途推定シートの折り返し・縦位置上揃え設定
                    if sheet_name == "資金移動用途推定":
                        col_name = cols[col_idx-1]
                        if col_name in ["資金移動日", "資金移動仕訳", "用途推定仕訳"]:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        else:
                            curr_align = cell.alignment
                            cell.alignment = Alignment(
                                horizontal=curr_align.horizontal if curr_align else "left",
                                vertical="top"
                            )
                        
                # 長期未回収売掛シートでの行背景色（評価に基づく）
                if sheet_name == "長期未回収売掛":
                    eval_val = str(row_data[cols.index("評価")]) if "評価" in cols else ""
                    origin_val = str(row_data[cols.index("発生原因")]) if "発生原因" in cols else ""
                    fill_to_apply = {
                        "grey": fill_grey,
                        "red": fill_red,
                        "orange": fill_orange,
                        "yellow": fill_yellow,
                    }.get(select_long_ar_fill_key(origin_val, eval_val))
                        
                    if fill_to_apply:
                        for cell in ws[row_idx]:
                            cell.fill = fill_to_apply
                            
        # 預金体力推移シートでの自動グラフ化（B/Sありの場合）
        if sheet_name == "預金体力推移" and bs_available and not df_sheet6.empty:
            chart = LineChart()
            chart.title = "預金体力推移（縦軸は「残高(円)」、横軸は「対象年月」）"
            chart.style = 10
            
            # グラフのサイズをやや広げて視認性と余白を確保
            chart.width = 20
            chart.height = 13
            
            # プロットエリア自体のレイアウトを設定して、上下左右に白い余白を確保
            # これにより、タイトル、軸タイトル、軸数値、凡例がグラフ線や枠線と被るのを防ぎます
            chart.plot_area.layout = Layout(
                manualLayout=ManualLayout(
                    x=0.15,      # 左余白 (全体を1.0とした割合、左から15%の位置から開始)
                    y=0.15,      # 上余白 (上から15%の位置から開始)
                    w=0.70,      # プロットエリアの幅 (右側に15%の凡例用余白を確保)
                    h=0.68,      # プロットエリアの高さ (下側に17%の横軸ラベル・タイトル用余白を確保)
                    xMode="edge",
                    yMode="edge"
                )
            )
            
            # 月末残高と月内最低残高 (B列とC列)
            data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(df_sheet6)+1)
            # 対象年月 (A列)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_sheet6)+1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # 軸の明示的表示（削除フラグをFalseに設定し、数値フォーマットを適用）
            chart.y_axis.delete = False
            chart.x_axis.delete = False
            chart.y_axis.number_format = '#,##0'
            
            # 凡例の位置調整：右上かつプロットエリア外に配置してグラフと重ねない
            chart.legend.position = "tr"  # Top Right (右上)
            chart.legend.overlay = False  # グラフエリアと重ねない
            
            # 横軸ラベル（年月）が重なるのを防ぐために-45度斜め回転
            from openpyxl.chart.text import RichText
            from openpyxl.drawing.text import RichTextProperties
            chart.x_axis.txPr = RichText(
                bodyPr=RichTextProperties(
                    rot="-2700000",  # -45度 (角度 * -60,000)
                    anchor="ctr",
                    anchorCtr="1",
                    spcFirstLastPara="1",
                    vertOverflow="ellipsis",
                    wrap="square"
                )
            )
            
            # グラフの折れ線のデザイン変更 (青基調、直線的)
            colors = ["1B365D", "4169E1"]
            for i, color_hex in enumerate(colors):
                if i < len(chart.series):
                    s = chart.series[i]
                    s.graphicalProperties.line.solidFill = color_hex
                    s.graphicalProperties.line.width = 25000  # 2.5pt
                    s.smooth = False
            
            # グラフの配置
            ws.add_chart(chart, "F2")
            
        # 列幅の設定
        if sheet_name == "資金移動用途推定":
            # 「資金移動用途推定」シートは自動調整を行わず、ファイル最上部で指定された固定幅を直接適用する
            for col_idx, col_name in enumerate(cols, 1):
                col_letter = get_column_letter(col_idx)
                width = BANK_LIST_COLUMN_WIDTHS.get(col_name, 15)
                ws.column_dimensions[col_letter].width = width
        else:
            # その他のシートは従来どおり自動スケーリング（改行考慮・全列合計最大200制限）を適用する
            col_widths = {}
            total_width = 0
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    # 改行で分割し、各行の中での最大長を測定する
                    lines = val_str.split('\n')
                    for line in lines:
                        length = sum(2 if ord(c) > 127 else 1 for c in line)
                        if length > max_len:
                            max_len = length
                
                # 推奨幅（文字数 + バッファ）
                recommended_width = max(max_len + 4, 12)
                col_widths[col_letter] = recommended_width
                total_width += recommended_width
                
            # 全列合計の列幅を最大200に制限するスケーリング調整
            MAX_TOTAL_WIDTH = 200
            if total_width > MAX_TOTAL_WIDTH:
                scale_factor = MAX_TOTAL_WIDTH / total_width
                for col_letter, w in col_widths.items():
                    col_widths[col_letter] = max(w * scale_factor, 8) # 最小幅は8を維持
                    
            # 列幅の適用
            for col_letter, w in col_widths.items():
                ws.column_dimensions[col_letter].width = w
            
    # デフォルトのSheetを削除
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    # bytes に変換
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()
    
    return excel_bytes, sales_index_data
